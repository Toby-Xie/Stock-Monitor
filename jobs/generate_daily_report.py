import os
from datetime import datetime
from io import BytesIO
from typing import Any, Optional

import pandas as pd
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import create_engine, text


API_BASE_URL = os.getenv("API_BASE_URL", "http://stock-monitor-api:8000/api")
REPORT_OUTPUT_DIR = os.getenv("REPORT_OUTPUT_DIR", "reports")

PG_URL = os.getenv(
    "DATABASE_URL",
    "postgresql+psycopg2://postgres:123456@db:5432/stockdb",
)

engine = create_engine(PG_URL, pool_pre_ping=True)


def get_json(endpoint: str, timeout: int = 120) -> dict[str, Any]:
    url = f"{API_BASE_URL.rstrip('/')}/{endpoint.lstrip('/')}"
    response = requests.get(url, timeout=timeout)
    response.raise_for_status()
    return response.json()


def payload_to_df(payload: dict) -> pd.DataFrame:
    """
    兼容这几种返回：
    1. {"data": [...]}
    2. {"data": {...}}
    3. [...]
    4. {...}
    """
    if isinstance(payload, list):
        return pd.DataFrame(payload)

    if not isinstance(payload, dict):
        return pd.DataFrame()

    data = payload.get("data", payload)

    if isinstance(data, list):
        return pd.DataFrame(data)

    if isinstance(data, dict):
        return pd.DataFrame([data])

    return pd.DataFrame()


def as_float(value, default=None):
    try:
        if value is None or pd.isna(value):
            return default
        return float(value)
    except Exception:
        return default


def load_market_pe_df() -> pd.DataFrame:
    payload = get_json("/market-pe/scan")
    return payload_to_df(payload)


def load_market_turn_df() -> pd.DataFrame:
    payload = get_json("/market-turn/scan")
    return payload_to_df(payload)


def load_margin_df() -> pd.DataFrame:
    payload = get_json("/margin")
    return payload_to_df(payload)


def calc_pe_percentile_10y(
    pe_df: pd.DataFrame,
) -> tuple[Optional[float], Optional[float], str]:
    """
    market_pe API 返回字段示例：
    {
        "trade_date": "2026-04-29",
        "market_pe": 23.156215,
        "matched_stock_count": 4740,
        "total_market_value": 126040249684576.16,
        "total_implied_profit": 5443041955094.03,
        ...
    }

    这里用 market_pe 历史序列计算当前 market_pe 在样本中的分位。
    """
    if pe_df.empty or "market_pe" not in pe_df.columns:
        return None, None, ""

    pe_df = pe_df.copy()
    pe_df["trade_date"] = pd.to_datetime(pe_df["trade_date"], errors="coerce")
    pe_df["market_pe"] = pd.to_numeric(pe_df["market_pe"], errors="coerce")

    pe_df = pe_df.dropna(subset=["trade_date", "market_pe"])
    pe_df = pe_df.sort_values("trade_date", ascending=False)

    if pe_df.empty:
        return None, None, ""

    latest = pe_df.iloc[0]
    current_pe = as_float(latest.get("market_pe"))

    hist = pe_df["market_pe"].dropna()

    percentile = None
    if current_pe is not None and not hist.empty:
        percentile = float((hist <= current_pe).mean() * 100)

    latest_date = latest["trade_date"].strftime("%Y-%m-%d")

    return current_pe, percentile, latest_date


def get_latest_market_turn(
    turn_df: pd.DataFrame,
) -> tuple[Optional[float], str]:
    """
    market_turn API 返回字段示例：
    {
        "trade_date": "2026-04-29",
        "market_turn": 2.51427,
        "matched_stock_count": 4692,
        "total_circulating_market_value": 125775787384071.45,
        "weighted_turn_value": 316234287642035.06,
        ...
    }
    """
    if turn_df.empty or "market_turn" not in turn_df.columns:
        return None, ""

    turn_df = turn_df.copy()
    turn_df["trade_date"] = pd.to_datetime(turn_df["trade_date"], errors="coerce")
    turn_df["market_turn"] = pd.to_numeric(turn_df["market_turn"], errors="coerce")

    turn_df = turn_df.dropna(subset=["trade_date", "market_turn"])
    turn_df = turn_df.sort_values("trade_date", ascending=False)

    if turn_df.empty:
        return None, ""

    latest = turn_df.iloc[0]
    market_turn = as_float(latest.get("market_turn"))
    latest_date = latest["trade_date"].strftime("%Y-%m-%d")

    return market_turn, latest_date


def get_latest_margin_balance(
    margin_df: pd.DataFrame,
) -> tuple[Optional[float], Optional[float], str]:
    """
    margin API 返回字段示例：
    {
        "trade_date": "2026-04-28",
        "exchange": "SSE",
        "fin_balance": "1373440361754.0",
        "fin_buy": "120599172093.0",
        "sec_volume": "2412675025.0",
        "sec_sell": "44328156.0",
        "sec_balance": "12943131485.0",
        "margin_balance": "1386383493239.0",
        ...
    }

    日报指标用“融资余额”，所以取 fin_balance。
    同一天可能有 SSE / SZSE 两行，因此按 trade_date 汇总。
    """
    required_cols = {"trade_date", "fin_balance"}
    if margin_df.empty or not required_cols.issubset(set(margin_df.columns)):
        return None, None, ""

    margin_df = margin_df.copy()
    margin_df["trade_date"] = pd.to_datetime(margin_df["trade_date"], errors="coerce")
    margin_df["fin_balance"] = pd.to_numeric(margin_df["fin_balance"], errors="coerce")

    margin_df = margin_df.dropna(subset=["trade_date", "fin_balance"])

    if margin_df.empty:
        return None, None, ""

    daily = (
        margin_df.groupby("trade_date", as_index=False)["fin_balance"]
        .sum()
        .sort_values("trade_date", ascending=False)
    )

    if daily.empty:
        return None, None, ""

    latest = daily.iloc[0]
    latest_balance = as_float(latest.get("fin_balance"))
    latest_date = latest["trade_date"].strftime("%Y-%m-%d")

    change_pct = None
    if len(daily) >= 2:
        prev = daily.iloc[1]
        prev_balance = as_float(prev.get("fin_balance"))

        if latest_balance is not None and prev_balance not in (None, 0):
            change_pct = (latest_balance - prev_balance) / prev_balance * 100

    return latest_balance, change_pct, latest_date


def build_indicator_table(
    market_pe_df: pd.DataFrame,
    market_turn_df: pd.DataFrame,
    margin_df: pd.DataFrame,
) -> pd.DataFrame:
    current_pe, pe_percentile, pe_date = calc_pe_percentile_10y(market_pe_df)
    market_turn, turn_date = get_latest_market_turn(market_turn_df)
    margin_balance, margin_change_pct, margin_date = get_latest_margin_balance(margin_df)

    pe_output = "" if pe_percentile is None else f"{pe_percentile:.1f}%"
    pe_logic = "" if pe_percentile is None else f"<40%? {'是' if pe_percentile < 40 else '否'}"

    turn_output = "" if market_turn is None else f"{market_turn:.2f}%"
    turn_logic = "" if market_turn is None else f">2.5%? {'是' if market_turn > 2.5 else '否'}"

    if margin_balance is None:
        margin_output = ""
    else:
        margin_output = f"{margin_balance / 1e8:.2f}亿"

    if margin_change_pct is None:
        margin_logic = "较上一交易日变化率：数据不足"
    else:
        margin_logic = f"较上一交易日变化率：{margin_change_pct:+.2f}%"

    rows = [
        {
            "指标": "中证全指 PE-TTM 分位（10年）",
            "计算/来源": "market_pe_daily.market_pe",
            "输出": pe_output,
            "判断逻辑": pe_logic,
            "数据日期": pe_date,
            "原始值": "" if current_pe is None else round(current_pe, 4),
        },
        {
            "指标": "全 A 换手率",
            "计算/来源": "market_turn_daily.market_turn",
            "输出": turn_output,
            "判断逻辑": turn_logic,
            "数据日期": turn_date,
            "原始值": "" if market_turn is None else round(market_turn, 4),
        },
        {
            "指标": "融资余额",
            "计算/来源": "market_margin_daily.fin_balance，按交易日汇总 SSE + SZSE",
            "输出": margin_output,
            "判断逻辑": margin_logic,
            "数据日期": margin_date,
            "原始值": "" if margin_balance is None else round(margin_balance, 2),
        },
    ]

    return pd.DataFrame(rows)


def load_industry_turnover_from_db() -> pd.DataFrame:
    """
    读取行业成交额表。

    industry_turnover_daily 字段：
    trade_date
    sw_l1_code
    sw_l1_name
    industry_amount
    total_amount
    industry_amount_ratio
    stock_count
    """
    sql = text("""
        SELECT
            trade_date,
            sw_l1_code,
            sw_l1_name,
            industry_amount,
            total_amount,
            industry_amount_ratio,
            stock_count
        FROM industry_turnover_daily
        ORDER BY sw_l1_code, trade_date
    """)

    with engine.connect() as conn:
        df = pd.read_sql(sql, conn)

    if df.empty:
        return df

    df["trade_date"] = pd.to_datetime(df["trade_date"], errors="coerce")

    for col in [
        "industry_amount",
        "total_amount",
        "industry_amount_ratio",
        "stock_count",
    ]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


def build_industry_crowding_table() -> pd.DataFrame:
    """
    行业拥挤度：
    - 每日拥挤度 = industry_amount_ratio
    - 20日拥挤度 = 按行业 rolling 20 日均值
    """
    raw = load_industry_turnover_from_db()

    columns = [
        "行业",
        "数据日期",
        "当日成交额占比",
        "拥挤度(20日)",
        "拥挤判断",
        "北向周净流入",
        "建议",
    ]

    if raw.empty:
        return pd.DataFrame(columns=columns)

    raw = raw.dropna(
        subset=["trade_date", "sw_l1_code", "sw_l1_name", "industry_amount_ratio"]
    ).copy()

    raw = raw.sort_values(["sw_l1_code", "trade_date"])

    if raw.empty:
        return pd.DataFrame(columns=columns)

    raw["crowding_20d"] = (
        raw.groupby("sw_l1_code")["industry_amount_ratio"]
        .transform(lambda s: s.rolling(20, min_periods=5).mean())
    )

    latest_date = raw["trade_date"].max()
    latest = raw[raw["trade_date"] == latest_date].copy()
    latest = latest.sort_values("crowding_20d", ascending=False)

    def crowding_judgement(value):
        if pd.isna(value):
            return "缺数据"
        if value >= 0.08:
            return "偏拥挤"
        if value <= 0.02:
            return "低拥挤"
        return "正常"

    def advice(value):
        if pd.isna(value):
            return ""
        if value >= 0.08:
            return "谨慎"
        if value <= 0.02:
            return "可关注"
        return "可关注"

    result = pd.DataFrame(
        {
            "行业": latest["sw_l1_name"],
            "数据日期": latest["trade_date"].dt.strftime("%Y-%m-%d"),
            "当日成交额占比": latest["industry_amount_ratio"].map(
                lambda x: "" if pd.isna(x) else f"{x * 100:.2f}%"
            ),
            "拥挤度(20日)": latest["crowding_20d"].map(
                lambda x: "" if pd.isna(x) else f"{x * 100:.2f}%"
            ),
            "拥挤判断": latest["crowding_20d"].map(crowding_judgement),
            "北向周净流入": "",
            "建议": latest["crowding_20d"].map(advice),
        }
    )

    return result.reset_index(drop=True)


def autosize_columns(worksheet, max_width: int = 40):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = cell.value
            if value is not None:
                max_length = max(max_length, len(str(value)))

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, max_width)


def style_sheet(worksheet):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    worksheet.freeze_panes = "A2"


def build_report_data() -> tuple[pd.DataFrame, pd.DataFrame]:
    market_pe_df = load_market_pe_df()
    market_turn_df = load_market_turn_df()
    margin_df = load_margin_df()

    indicator_df = build_indicator_table(
        market_pe_df=market_pe_df,
        market_turn_df=market_turn_df,
        margin_df=margin_df,
    )

    industry_df = build_industry_crowding_table()

    return indicator_df, industry_df


def write_report_to_excel(
    output,
    indicator_df: pd.DataFrame,
    industry_df: pd.DataFrame,
):
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        indicator_df.to_excel(writer, index=False, sheet_name="market_indicators")
        industry_df.to_excel(writer, index=False, sheet_name="industry_crowding")

        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            autosize_columns(worksheet)
            style_sheet(worksheet)


def generate_daily_report(
    output_dir: str = REPORT_OUTPUT_DIR,
    report_date: Optional[str] = None,
) -> str:
    """
    生成 Excel 文件并保存到本地目录。
    返回文件路径。
    """
    if report_date is None:
        report_date = datetime.now().strftime("%Y-%m-%d")

    os.makedirs(output_dir, exist_ok=True)

    indicator_df, industry_df = build_report_data()

    output_path = os.path.join(output_dir, f"daily_report_{report_date}.xlsx")

    write_report_to_excel(output_path, indicator_df, industry_df)

    print(f"daily report generated: {output_path}")
    return output_path


def generate_daily_report_bytes(
    report_date: Optional[str] = None,
) -> BytesIO:
    """
    生成 Excel 到内存，给 FastAPI StreamingResponse 下载用。
    """
    if report_date is None:
        report_date = datetime.now().strftime("%Y-%m-%d")

    indicator_df, industry_df = build_report_data()

    output = BytesIO()
    write_report_to_excel(output, indicator_df, industry_df)
    output.seek(0)

    return output


if __name__ == "__main__":
    path = generate_daily_report()
    print(path)